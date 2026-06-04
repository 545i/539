"""excel_report 模組測試:驗證活頁簿工作表、bytes 內容與可重新載入。"""
from __future__ import annotations

import io

import openpyxl
import pytest

from core import backtest, excel_report, loader


@pytest.fixture
def sample_df():
    """60 期固定 seed 範例資料。"""
    return loader.generate_sample(60)


def test_build_workbook_has_expected_sheets(sample_df):
    """不傳 backtest_results 時,應含基本四張工作表。"""
    wb = excel_report.build_workbook(sample_df)
    names = wb.sheetnames
    for expected in ["免責聲明", "開獎資料", "號碼頻率", "凱莉分析"]:
        assert expected in names
    # 未傳回測結果則不應有該表
    assert "回測結果" not in names


def test_draw_sheet_content(sample_df):
    """開獎資料表頭與列數正確,日期為 YYYY-MM-DD 字串。"""
    wb = excel_report.build_workbook(sample_df)
    ws = wb["開獎資料"]
    assert [ws.cell(row=1, column=c).value for c in range(1, 7)] == \
        ["日期", "號1", "號2", "號3", "號4", "號5"]
    # 表頭 1 列 + 60 期資料
    assert ws.max_row == 61
    # 日期格式檢查
    val = ws.cell(row=2, column=1).value
    assert isinstance(val, str) and len(val) == 10 and val[4] == "-"


def test_frequency_sheet_has_chart(sample_df):
    """號碼頻率表含 39 列號碼與一張 BarChart。"""
    wb = excel_report.build_workbook(sample_df)
    ws = wb["號碼頻率"]
    assert ws.cell(row=1, column=1).value == "號碼"
    assert ws.cell(row=1, column=2).value == "出現次數"
    # 1~39 共 39 個號碼
    assert ws.max_row == 40
    assert len(ws._charts) == 1


def test_build_report_bytes_is_xlsx(sample_df):
    """build_report_bytes 回傳非空 bytes 且為 zip(xlsx)開頭 b'PK'。"""
    data = excel_report.build_report_bytes(sample_df)
    assert isinstance(data, bytes)
    assert len(data) > 0
    assert data[:2] == b"PK"


def test_bytes_can_be_reloaded(sample_df):
    """產出的 bytes 可被 openpyxl 重新載入。"""
    data = excel_report.build_report_bytes(sample_df)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "免責聲明" in wb.sheetnames
    assert "號碼頻率" in wb.sheetnames


def test_backtest_results_adds_sheet(sample_df):
    """傳入 backtest_results 時多出「回測結果」工作表並含圖表。"""
    results = backtest.compare(sample_df, ["random", "hot"], start_index=30)
    wb = excel_report.build_workbook(sample_df, backtest_results=results)
    assert "回測結果" in wb.sheetnames
    ws = wb["回測結果"]
    assert ws.cell(row=1, column=1).value == "策略"
    assert ws.cell(row=1, column=9).value == "報酬率%"
    # 兩個策略 + 表頭
    assert ws.max_row == 3
    assert len(ws._charts) == 1


def test_save_report_writes_file(sample_df, tmp_path):
    """save_report 寫出可載入的實體檔。"""
    out = tmp_path / "report.xlsx"
    excel_report.save_report(sample_df, out)
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert "凱莉分析" in wb.sheetnames
