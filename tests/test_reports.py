"""排行榜 / 匯出 / 設定(開獎更新狀態)三個端點。

跟 test_ledger 一樣自己組 app,不掛 backend.main —— 驗的是端點本身。
ledger.db 導到 tmp_path;「立即抓取」把 catch_up 換掉,測試不連外網。
"""
from __future__ import annotations

import io

import pytest

pytest.importorskip("httpx", reason="TestClient 需要 httpx")

import openpyxl  # noqa: E402
from fastapi import FastAPI  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

from backend import ledger_store  # noqa: E402
from backend.routers import export, leaderboard, settings  # noqa: E402
from core import auth, autoupdate  # noqa: E402

P = "/api"


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(ledger_store, "_db_path", lambda: tmp_path / "ledger.db")
    app = FastAPI()
    for r in (leaderboard.router, export.router, settings.router):
        app.include_router(r, prefix=P)
    return TestClient(app)


def hdr(name: str) -> dict:
    return {"Authorization": f"Bearer {auth.make_token(name)}"}


@pytest.fixture()
def seeded():
    """alice 賺 100(兩勝一敗)、bob 賠 50。"""
    ledger_store.add_entry("alice", "single", {"pnl": 200, "cost": 300, "payout": 500})
    ledger_store.add_entry("alice", "single", {"pnl": -100, "cost": 100, "payout": 0})
    ledger_store.add_entry("alice", "combo", {"pnl": 0, "cost": 0, "payout": 0})
    ledger_store.add_entry("bob", "pillar1800", {"pnl": -50, "cost": 50, "payout": 0})


# ── 排行榜 ────────────────────────────────────────────────
def test_leaderboard_requires_login(client):
    assert client.get(f"{P}/leaderboard").status_code == 401


def test_leaderboard_ranks_by_pnl(client, seeded):
    body = client.get(f"{P}/leaderboard", headers=hdr("bob")).json()
    assert body["me"] == "bob"
    users = body["users"]
    assert [u["username"] for u in users] == ["alice", "bob"]
    assert [u["rank"] for u in users] == [1, 2]

    alice = users[0]
    assert alice["total_pnl"] == 100 and alice["rounds"] == 3
    assert alice["wins"] == 1 and alice["losses"] == 1
    assert alice["total_cost"] == 400 and alice["roi"] == pytest.approx(0.25)
    assert alice["is_me"] is False
    assert users[1]["is_me"] is True


def test_leaderboard_roi_none_when_no_cost(client):
    ledger_store.add_entry("zoe", "multi", {"pnl": 0})
    body = client.get(f"{P}/leaderboard", headers=hdr("zoe")).json()
    assert body["users"][0]["roi"] is None


def test_leaderboard_modes_aggregate(client, seeded):
    body = client.get(f"{P}/leaderboard", headers=hdr("alice")).json()
    modes = {m["mode"]: m for m in body["modes"]}
    assert modes["single"]["rounds"] == 2 and modes["single"]["total_pnl"] == 100
    assert modes["single"]["name"] == "單顆下注"
    assert modes["pillar1800"]["total_pnl"] == -50


def test_leaderboard_empty(client):
    body = client.get(f"{P}/leaderboard", headers=hdr("alice")).json()
    assert body["users"] == [] and body["modes"] == []


# ── 匯出 ──────────────────────────────────────────────────
def test_export_report_xlsx(client):
    r = client.get(f"{P}/export/report.xlsx?game=lotto539&limit=30")
    assert r.status_code == 200
    assert r.content[:2] == b"PK"
    assert "attachment" in r.headers["content-disposition"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert {"免責聲明", "開獎資料", "號碼頻率", "凱莉分析"} <= set(wb.sheetnames)
    assert wb["開獎資料"].max_row == 31          # 表頭 + 30 期
    assert wb["號碼頻率"].max_row == 40          # 表頭 + 39 號


def test_export_report_marksix_has_six_numbers(client):
    """六合彩 49 選 6:開獎資料要有 6 欄號碼、頻率表要有 49 個號。"""
    r = client.get(f"{P}/export/report.xlsx?game=marksix&limit=10")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["開獎資料"]
    assert [ws.cell(row=1, column=c).value for c in range(1, 8)] == \
        ["日期", "號1", "號2", "號3", "號4", "號5", "號6"]
    assert wb["號碼頻率"].max_row == 50


def test_export_report_unknown_game(client):
    assert client.get(f"{P}/export/report.xlsx?game=nope").status_code == 404


def test_export_ledger_xlsx(client, seeded):
    assert client.get(f"{P}/export/ledger.xlsx").status_code == 401
    r = client.get(f"{P}/export/ledger.xlsx", headers=hdr("alice"))
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "全部" in wb.sheetnames and "單顆下注" in wb.sheetnames
    ws = wb["全部"]
    assert ws.max_row == 4                       # 表頭 + alice 的 3 筆(不含 bob)
    assert ws.cell(row=2, column=15).value == 200     # 本局損益
    assert ws.cell(row=3, column=16).value == 100     # 累積損益
    assert wb["單顆下注"].max_row == 3


def test_export_ledger_json(client, seeded):
    r = client.get(f"{P}/export/ledger.json", headers=hdr("bob"))
    assert r.status_code == 200
    body = r.json()
    assert body["username"] == "bob" and len(body["entries"]) == 1
    assert body["entries"][0]["record"]["pnl"] == -50


# ── 設定:開獎更新狀態 ─────────────────────────────────────
def test_autoupdate_status_shape(client):
    body = client.get(f"{P}/settings/autoupdate").json()
    assert isinstance(body["scheduler_running"], bool)
    assert {g["key"] for g in body["games"]} == {"lotto539", "fantasy5", "marksix"}
    for g in body["games"]:
        assert g["scheduled"] is True
        assert g["latest"] and g["target"]
        assert isinstance(g["stale"], bool)
        assert set(g["status"]) == {"running", "msg", "error", "attempts",
                                    "added", "done_at"}


def test_fetch_now_requires_login(client):
    assert client.post(f"{P}/settings/fetch-now", json={}).status_code == 401


def test_fetch_now_single_game(client, monkeypatch):
    calls = []

    def fake(game_key, path):
        calls.append(game_key)
        import datetime as dt
        return {"fetched": 5, "added": 1, "latest": dt.date(2026, 8, 19)}

    monkeypatch.setattr(autoupdate, "catch_up", fake)
    r = client.post(f"{P}/settings/fetch-now", json={"game": "lotto539"},
                    headers=hdr("alice"))
    assert calls == ["lotto539"]
    row = r.json()["results"][0]
    assert row["ok"] and row["added"] == 1 and row["latest"] == "2026-08-19"


def test_fetch_now_all_games_survives_failure(client, monkeypatch):
    def fake(game_key, path):
        if game_key == "marksix":
            raise RuntimeError("來源站掛了")
        import datetime as dt
        return {"fetched": 1, "added": 0, "latest": dt.date(2026, 8, 19)}

    monkeypatch.setattr(autoupdate, "catch_up", fake)
    r = client.post(f"{P}/settings/fetch-now", json={}, headers=hdr("alice"))
    rows = {x["key"]: x for x in r.json()["results"]}
    assert len(rows) == 3
    assert rows["lotto539"]["ok"] is True
    assert rows["marksix"]["ok"] is False and "來源站掛了" in rows["marksix"]["error"]
