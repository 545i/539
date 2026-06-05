"""Excel 圖表報表模組:用 openpyxl 產生含「原生 Excel 圖表」的 .xlsx 報表。

提供三個對外函式:
- build_workbook(df, ...) -> openpyxl.Workbook  建立含多張工作表(部分綁定原生 BarChart)的活頁簿。
- build_report_bytes(df, **kwargs) -> bytes      把活頁簿存到記憶體 BytesIO 後回傳 bytes。
- save_report(df, path, **kwargs) -> None         把活頁簿寫到實體檔案。

只依賴 openpyxl / io / core.*,不碰網路。所有圖表皆為 Excel 原生圖表(非圖片)。
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference

from core import constants, stats


# ── 內部工具:寫表頭 ──────────────────────────────────────
def _write_header(ws, headers: list[str], row: int = 1) -> None:
    """在指定列寫入表頭文字。"""
    for col, text in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=text)


# ── 1. 免責聲明 ──────────────────────────────────────────
def _build_disclaimer_sheet(wb, game=None) -> None:
    """把免責聲明逐行寫入「免責聲明」工作表;有 game 時加註該遊戲資訊。"""
    ws = wb.active
    ws.title = "免責聲明"
    lines = []
    if game is not None:
        lines.append(f"遊戲:{game.name}(票價 {game.currency}{game.ticket_price:g})")
        lines.append(f"期望報酬率約 {game.expected_return():.2%}")
        lines.append(f"獎金結構:{game.prize_note}")
        lines.append(f"資料來源:{game.source_note}")
        lines.append("")
    lines.extend(constants.DISCLAIMER.split("\n"))
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)


# ── 2. 開獎資料 ──────────────────────────────────────────
def _build_draw_sheet(wb, df) -> None:
    """寫入每期開獎號碼;日期轉成 YYYY-MM-DD 字串。"""
    ws = wb.create_sheet("開獎資料")
    _write_header(ws, ["日期", "號1", "號2", "號3", "號4", "號5"])
    import pandas as pd  # 局部匯入,避免模組層相依過廣

    for r, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=r, column=1, value=pd.to_datetime(row["date"]).strftime("%Y-%m-%d"))
        for c in range(1, 6):
            ws.cell(row=r, column=1 + c, value=int(row[f"n{c}"]))


# ── 3. 號碼頻率(含 BarChart)─────────────────────────────
def _build_frequency_sheet(wb, df, freq) -> None:
    """寫入號碼出現次數,並綁定 BarChart 呈現分布。"""
    ws = wb.create_sheet("號碼頻率")
    _write_header(ws, ["號碼", "出現次數"])

    if freq is None:
        freq = stats.frequency(df)

    items = sorted(freq.items())  # 依號碼由小到大排序
    for r, (num, cnt) in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=int(num))
        ws.cell(row=r, column=2, value=int(cnt))

    last_row = 1 + len(items)
    chart = BarChart()
    chart.type = "col"
    chart.title = "號碼出現次數"
    chart.x_axis.title = "號碼"
    chart.y_axis.title = "出現次數"
    # 資料範圍含表頭(出現次數欄),類別為號碼欄
    data = Reference(ws, min_col=2, min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # 圖表放資料右側空白儲存格(第 4 欄)
    ws.add_chart(chart, "D2")


# ── 4. 回測結果(含 BarChart)─────────────────────────────
def _build_backtest_sheet(wb, backtest_results) -> None:
    """寫入各策略回測結果並以 BarChart 比較報酬率%。"""
    ws = wb.create_sheet("回測結果")
    _write_header(ws, ["策略", "期數", "中2", "中3", "中4", "中5",
                       "總投注", "總回收", "報酬率%"])

    for r, res in enumerate(backtest_results, start=2):
        hit = res.hit_dist or {}
        ws.cell(row=r, column=1, value=res.strategy)
        ws.cell(row=r, column=2, value=int(res.periods))
        ws.cell(row=r, column=3, value=int(hit.get(2, 0)))
        ws.cell(row=r, column=4, value=int(hit.get(3, 0)))
        ws.cell(row=r, column=5, value=int(hit.get(4, 0)))
        ws.cell(row=r, column=6, value=int(hit.get(5, 0)))
        ws.cell(row=r, column=7, value=round(float(res.total_bet), 2))
        ws.cell(row=r, column=8, value=round(float(res.total_return), 2))
        ws.cell(row=r, column=9, value=round(res.roi * 100, 4))

    last_row = 1 + len(backtest_results)
    chart = BarChart()
    chart.type = "col"
    chart.title = "各策略報酬率%"
    chart.x_axis.title = "策略"
    chart.y_axis.title = "報酬率%"
    # 報酬率% 在第 9 欄,類別(策略名稱)在第 1 欄
    data = Reference(ws, min_col=9, min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # 圖表放資料右側空白儲存格(第 11 欄 = K)
    ws.add_chart(chart, "K2")


# ── 5. 凱莉分析 ──────────────────────────────────────────
def _resolve_kelly_result(kelly_result):
    """取得凱莉分析結果;為 None 時嘗試呼叫 core.kelly.analyze_539()。

    若 kelly 模組尚未存在,退回以 constants 計算的基本期望值字典,
    確保「凱莉分析」工作表一定能產生。
    """
    if kelly_result is not None:
        return kelly_result
    try:
        from core import kelly  # 延遲匯入,避免硬相依
        return kelly.analyze_539()
    except Exception:
        # 退回基本期望值資訊(今彩539 期望報酬為負,凱莉比例為 0,不應投注)
        return {
            "每注期望獎金": round(constants.expected_prize(), 4),
            "每注成本": constants.TICKET_PRICE,
            "期望報酬率": round(constants.EXPECTED_RETURN, 6),
            "凱莉建議下注比例": 0.0,
            "結論": "期望值為負,凱莉準則建議下注比例為 0(不應投注)。",
        }


def _flatten(result, prefix: str = "") -> list[tuple[str, object]]:
    """把巢狀的凱莉結果攤平成 (項目, 數值) 的清單。"""
    rows: list[tuple[str, object]] = []
    if isinstance(result, dict):
        items = result.items()
    elif hasattr(result, "__dict__"):
        items = vars(result).items()
    else:
        return [(prefix or "數值", result)]

    for key, val in items:
        name = f"{prefix}{key}"
        if isinstance(val, dict) or hasattr(val, "__dict__"):
            rows.extend(_flatten(val, prefix=f"{name}."))
        elif isinstance(val, (list, tuple)):
            rows.append((name, ", ".join(str(v) for v in val)))
        else:
            rows.append((name, val))
    return rows


def _build_kelly_sheet(wb, kelly_result) -> None:
    """把凱莉分析欄位攤平成「項目/數值」兩欄。"""
    ws = wb.create_sheet("凱莉分析")
    _write_header(ws, ["項目", "數值"])
    result = _resolve_kelly_result(kelly_result)
    for r, (name, val) in enumerate(_flatten(result), start=2):
        ws.cell(row=r, column=1, value=str(name))
        ws.cell(row=r, column=2, value=val)


# ── 對外主函式 ───────────────────────────────────────────
def build_workbook(df, *, freq=None, backtest_results=None, kelly_result=None, game=None):
    """建立含原生 Excel 圖表的活頁簿並回傳(不存檔)。

    工作表順序:免責聲明、開獎資料、號碼頻率、(回測結果)、凱莉分析。
    backtest_results 為 None 或空時略過「回測結果」工作表。
    game 為 GameConfig 時,免責聲明會加註該遊戲的票價/期望報酬/資料來源。
    """
    wb = openpyxl.Workbook()
    _build_disclaimer_sheet(wb, game)
    _build_draw_sheet(wb, df)
    _build_frequency_sheet(wb, df, freq)
    if backtest_results:
        _build_backtest_sheet(wb, backtest_results)
    _build_kelly_sheet(wb, kelly_result)
    return wb


def build_report_bytes(df, **kwargs) -> bytes:
    """建立報表並回傳 .xlsx 的位元組內容(存到記憶體 BytesIO)。"""
    wb = build_workbook(df, **kwargs)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_report(df, path, **kwargs) -> None:
    """建立報表並寫入指定路徑的 .xlsx 檔案。"""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(df, **kwargs)
    wb.save(path)
